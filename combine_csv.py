#!/usr/bin/env python3
"""
Combine all CSV accessibility results into a single Excel file.
"""
import os
import csv
import glob
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

def read_csv_file(filepath):
    """Read a CSV file and return metadata and violation rows."""
    metadata = {}
    rows = []
    with open(filepath, 'r', encoding='utf-8') as f:
        reader = csv.reader(f)
        header_found = False
        current_section = None
        
        for row in reader:
            # Skip metadata lines (starting with #)
            if row and row[0].strip().startswith('#'):
                continue
            # Skip empty rows
            if not any(cell.strip() for cell in row):
                continue
            
            # Extract metadata from Test Information section
            if len(row) > 0:
                first_cell = row[0].strip()
                
                # Check for section headers
                if first_cell == 'Test Information':
                    current_section = 'test'
                    continue
                elif first_cell == 'Environment Information':
                    current_section = 'environment'
                    continue
                elif first_cell == 'Test Results':
                    current_section = 'results'
                    continue
                
                # Extract metadata values
                if current_section == 'test' and len(row) >= 2:
                    key = first_cell
                    value = row[1].strip() if len(row) > 1 else ''
                    if key == 'Test URL':
                        metadata['url'] = value
                    elif key == 'Timestamp':
                        metadata['timestamp'] = value
                    elif key == 'Test Engine':
                        metadata['test_engine'] = value
                    elif key == 'Test Runner':
                        metadata['test_runner'] = value
                
                elif current_section == 'environment' and len(row) >= 2:
                    key = first_cell
                    value = row[1].strip() if len(row) > 1 else ''
                    if key == 'User Agent':
                        metadata['user_agent'] = value
                    elif key == 'Window Size':
                        metadata['window_size'] = value
                    elif key == 'Orientation':
                        metadata['orientation'] = value
            
            # Look for the actual header row that contains "Rule ID"
            if not header_found:
                if 'Rule ID' in row:
                    header_found = True
                    rows.append(row)  # Include the header row
                # Skip all rows until we find the header
                continue
            # After header is found, include all data rows
            rows.append(row)
    
    return metadata, rows

def sanitize_sheet_name(name):
    """Sanitize sheet name to be Excel-compatible (max 31 chars, no invalid chars)."""
    # Excel sheet name restrictions: max 31 chars, no: \ / ? * [ ]
    invalid_chars = ['\\', '/', '?', '*', '[', ']']
    for char in invalid_chars:
        name = name.replace(char, '_')
    # Truncate to 31 characters
    return name[:31]

def extract_url_path_from_filename(filename):
    """Extract a meaningful URL path from the CSV filename."""
    # Remove common prefixes
    name = filename.replace('accessibility-results-', '').replace('.csv', '')
    
    # Extract domain and path
    # Format: domain-com-path or domain-com-path-path2
    parts = name.split('-')
    
    # Find where the domain ends (usually after 'com' or 'net' etc)
    domain_end_idx = None
    for i, part in enumerate(parts):
        if part in ['com', 'net', 'org', 'io', 'co']:
            domain_end_idx = i + 1
            break
    
    if domain_end_idx and domain_end_idx < len(parts):
        # Get the path parts after domain
        path_parts = parts[domain_end_idx:]
        if path_parts:
            # Join path parts and limit length
            path = '-'.join(path_parts)
            # If too long, take meaningful parts
            if len(path) > 28:
                # Try to keep the last meaningful segment
                if len(path_parts) > 1:
                    # Take last 2-3 parts
                    path = '-'.join(path_parts[-3:])
                else:
                    path = path[-28:]
            return path if path else 'root'
    
    # Fallback: use 'root' for home page, or last meaningful part
    if len(parts) <= 2:  # Just domain, no path
        return 'root'
    return parts[-1][:31] if parts else 'page'

def combine_csv_files(results_dir='results', output_file='combined_accessibility_results.xlsx'):
    """Combine all CSV files into a single Excel file with separate sheets per URL."""
    csv_files = glob.glob(os.path.join(results_dir, '*.csv'))
    
    # Exclude generated files from being combined
    excluded_files = ['all_violations_except_nested_interactive.csv', 'combined_accessibility_results.xlsx']
    csv_files = [f for f in csv_files if os.path.basename(f) not in excluded_files]
    
    if not csv_files:
        print(f"No CSV files found in {results_dir}")
        return
    
    print(f"Found {len(csv_files)} CSV file(s) to combine")
    
    # Create a new workbook
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)
    
    # Header style
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    total_rows = 0
    sheets_created = 0
    
    # Process each CSV file into its own sheet
    for csv_file in sorted(csv_files):
        filename = os.path.basename(csv_file)
        print(f"Processing {filename}...")
        
        metadata, rows = read_csv_file(csv_file)
        
        if not rows or len(rows) < 2:  # Need at least header + 1 data row
            print(f"  Skipping {filename} - no violation data found")
            continue
        
        # Create sheet name from filename (extract meaningful URL path)
        sheet_name = sanitize_sheet_name(extract_url_path_from_filename(filename))
        
        # Ensure unique sheet names (Excel doesn't allow duplicates)
        original_sheet_name = sheet_name
        counter = 1
        while sheet_name in [s.title for s in wb.worksheets]:
            sheet_name = sanitize_sheet_name(f"{original_sheet_name[:28]}_{counter}")
            counter += 1
        
        # Create a new worksheet for this URL
        ws = wb.create_sheet(title=sheet_name)
        
        # Metadata style
        metadata_label_font = Font(bold=True, size=11)
        metadata_value_font = Font(size=10)
        
        current_row = 1
        
        # Write metadata section at the top
        if metadata:
            # Metadata header
            ws.merge_cells(f'A{current_row}:B{current_row}')
            metadata_header_cell = ws.cell(row=current_row, column=1, value='Test Information')
            metadata_header_cell.font = Font(bold=True, size=12)
            metadata_header_cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
            current_row += 1
            
            # Write metadata fields
            if metadata.get('url'):
                ws.cell(row=current_row, column=1, value='Test URL:').font = metadata_label_font
                ws.cell(row=current_row, column=2, value=metadata['url']).font = metadata_value_font
                current_row += 1
            
            if metadata.get('timestamp'):
                ws.cell(row=current_row, column=1, value='Timestamp:').font = metadata_label_font
                ws.cell(row=current_row, column=2, value=metadata['timestamp']).font = metadata_value_font
                current_row += 1
            
            if metadata.get('test_engine'):
                ws.cell(row=current_row, column=1, value='Test Engine:').font = metadata_label_font
                ws.cell(row=current_row, column=2, value=metadata['test_engine']).font = metadata_value_font
                current_row += 1
            
            if metadata.get('test_runner'):
                ws.cell(row=current_row, column=1, value='Test Runner:').font = metadata_label_font
                ws.cell(row=current_row, column=2, value=metadata['test_runner']).font = metadata_value_font
                current_row += 1
            
            if metadata.get('window_size'):
                ws.cell(row=current_row, column=1, value='Window Size:').font = metadata_label_font
                ws.cell(row=current_row, column=2, value=metadata['window_size']).font = metadata_value_font
                current_row += 1
            
            # Add spacing before violation data
            current_row += 1
        
        header = rows[0]
        data_rows = rows[1:]
        
        # Write violation header row
        violation_header_row = current_row
        for col_idx, value in enumerate(header, start=1):
            cell = ws.cell(row=current_row, column=col_idx, value=value)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        current_row += 1
        
        # Write violation data rows
        for row in data_rows:
            for col_idx, value in enumerate(row, start=1):
                if col_idx <= len(header):  # Ensure we don't exceed header columns
                    cell = ws.cell(row=current_row, column=col_idx, value=value)
                    cell.alignment = Alignment(vertical='top', wrap_text=True)
            current_row += 1
        
        # Auto-adjust column widths for this sheet
        # Adjust metadata columns
        ws.column_dimensions['A'].width = 15  # Label column
        ws.column_dimensions['B'].width = 60  # Value column (wider for URLs)
        
        # Adjust violation data columns
        for col_idx in range(1, len(header) + 1):
            column_letter = get_column_letter(col_idx)
            max_length = 0
            
            # Only check rows from violation header onwards
            for row_num in range(violation_header_row, current_row + 1):
                cell = ws.cell(row=row_num, column=col_idx)
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            ws.column_dimensions[column_letter].width = max(adjusted_width, 10)  # Min width 10
        
        sheets_created += 1
        total_rows += len(data_rows)
        print(f"  Created sheet '{sheet_name}' with {len(data_rows)} violation(s)")
    
    if sheets_created == 0:
        print("No sheets created - no violation data found in any CSV files")
        return
    
    # Save the workbook
    output_path = os.path.join(results_dir, output_file)
    wb.save(output_path)
    print(f"\nCombined results saved to: {output_path}")
    print(f"Total sheets: {sheets_created}")
    print(f"Total violation rows: {total_rows}")

if __name__ == '__main__':
    import sys
    results_dir = sys.argv[1] if len(sys.argv) > 1 else 'results'
    combine_csv_files(results_dir)
